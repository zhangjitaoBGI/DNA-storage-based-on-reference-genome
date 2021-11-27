# -*- coding: utf-8 -*-
"""
Created on Fri Jul 23 11:44:07 2021

@author: Jitao Zhang
"""
from openpyxl import load_workbook
import pandas as pd

if __name__ == '__main__':
    '''
    writer = pd.ExcelWriter("./jpeg_result.xlsx",engine="openpyxl")
    wb = load_workbook(writer.path)
    writer.book = wb
    df = pd.DataFrame({"a":[0],'v':[2]})
    df.to_excel(writer,sheet_name="sad")
    writer.save()
    writer.close()
    '''
    writer = pd.ExcelWriter("./jpeg_result.xlsx",engine='openpyxl')#file_path为自己需要保存到的路径
    book = load_workbook(writer.path)
    writer.book = book
    df = pd.DataFrame({"a":[0],'v':[2]})
    df.to_excel(excel_writer=writer,sheet_name="aaa")#aaa为你自己命名的sheet名
    writer.save()
