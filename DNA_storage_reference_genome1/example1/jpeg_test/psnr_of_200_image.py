# -*- coding: utf-8 -*-
"""
Created on Tue Sep  8 23:36:55 2020

@author: Jitao Zhang
"""
import cv2
import numpy as np
import math
import os
from openpyxl import load_workbook,Workbook
import pandas as pd


def psnr(img1, img2):
   mse = np.mean( (img1/255. - img2/255.) ** 2 )
   if mse < 1.0e-10:
      return 100
   PIXEL_MAX = 1
   return 20 * math.log10(PIXEL_MAX / math.sqrt(mse))

def calculate_psnr(img1, img2):
    # img1 and img2 have range [0, 255]
    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    mse = np.mean((img1 - img2)**2)
    if mse == 0:
        return float('inf')
    return 20 * math.log10(255.0 / math.sqrt(mse))

def ssim(img1, img2):
    C1 = (0.01 * 255)**2
    C2 = (0.03 * 255)**2

    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    kernel = cv2.getGaussianKernel(11, 1.5)
    window = np.outer(kernel, kernel.transpose())

    mu1 = cv2.filter2D(img1, -1, window)[5:-5, 5:-5]  # valid
    mu2 = cv2.filter2D(img2, -1, window)[5:-5, 5:-5]
    mu1_sq = mu1**2
    mu2_sq = mu2**2
    mu1_mu2 = mu1 * mu2
    sigma1_sq = cv2.filter2D(img1**2, -1, window)[5:-5, 5:-5] - mu1_sq
    sigma2_sq = cv2.filter2D(img2**2, -1, window)[5:-5, 5:-5] - mu2_sq
    sigma12 = cv2.filter2D(img1 * img2, -1, window)[5:-5, 5:-5] - mu1_mu2

    ssim_map = ((2 * mu1_mu2 + C1) * (2 * sigma12 + C2)) / ((mu1_sq + mu2_sq + C1) *
                                                            (sigma1_sq + sigma2_sq + C2))
    return ssim_map.mean()

def calculate_ssim(img1, img2):
   '''calculate SSIM
   the same outputs as MATLAB's
   img1, img2: [0, 255]
   '''
   if not img1.shape == img2.shape:
       raise ValueError('Input images must have the same dimensions.')
   if img1.ndim == 2:
       return ssim(img1, img2)
   elif img1.ndim == 3:
       if img1.shape[2] == 3:
           ssims = []
           for i in range(3):
               ssims.append(ssim(img1, img2))
           return np.array(ssims).mean()
       elif img1.shape[2] == 1:
           return ssim(np.squeeze(img1), np.squeeze(img2))
   else:
       raise ValueError('Wrong input image dimensions.')
  


if __name__ == '__main__':
    '''
    img_monarch_512 = cv2.imread(r"E:\work in CNGB\DNA_storage\image_processing\Image_processing\Bmp\lena512_16_16_matrix\kl-test\image-result\dataset\monarch-512.png")

    img_lena_256_pred_x2 = cv2.imread(r"E:\work in CNGB\interim report\image\512\ESRGAN\lena-512-color24-244-256X2.png")
    img_lena_256_x2_e =  cv2.imread(r"E:\work in CNGB\interim report\image\512\ESRGAN\out_lena256-color-24-244_rlt_x2.png")
    img_monarch_256_pred_x2_e = cv2.imread(r"E:\work in CNGB\DNA_storage\image_processing\Image_processing\Bmp\lena512_16_16_matrix\kl-test\image-result\scale\train_100\monarch-244_rlt_x2_test.png")
    
    img_lena_512 = cv2.imread(r"E:\work in CNGB\ESRGAN\lena512_color_24.bmp")
    img_lena_256_pred_x2 = cv2.imread(r"E:\work in CNGB\ESRGAN\out_lena256-color24-244_rlt_prec_x2.bmp")
    img_lena_256_x2_e = cv2.imread(r"E:\work in CNGB\ESRGAN\out_lena256-color-24-244_rlt_x2.png")
    
    print(calculate_psnr(img_lena_512,img_lena_256_x2_e),calculate_ssim(img_lena_512,img_lena_256_x2_e))
    print(calculate_psnr(img_lena_512,img_lena_256_pred_x2),calculate_ssim(img_lena_512,img_lena_256_pred_x2))
    '''
    length = 543
    orginal_image_folder_path =  r"E:\work in CNGB\DNA_storage\image_processing\Image_processing\Bmp\lena512_16_16_matrix\datasets\DIV2K_train_HR\small-128\bmp"
    processed_image_folder_path = r"E:\work in CNGB\DNA_storage\DNA_storage_reference_genome1\example1\result\jpeg_compression_test\Div2k\length_{}".format(length)
    orginal_image_list = os.listdir(orginal_image_folder_path)[:200]

    image_psnr_dict = {"image":[],"PSNR":[],"SSIM":[]}
    
    for image in orginal_image_list:
        img1 = cv2.imread(orginal_image_folder_path+"/"+image)
        psnr_list = []
        ssim_list = []
        processed_img_path = processed_image_folder_path + "/" +image[:-4]
        processed_image_list = os.listdir(processed_img_path)
        for random_image in processed_image_list:
            random_image_path = processed_img_path + "/" + random_image
            img2 = cv2.imread(random_image_path)
            psnr_list.append(calculate_psnr(img1,img2))
            ssim_list.append(calculate_ssim(img1,img2))
        image_psnr_dict["image"].append(image[:-4]+".jpg")        
        image_psnr_dict["PSNR"].append(np.mean(psnr_list))
        image_psnr_dict["SSIM"].append(np.mean(ssim_list))
        print("Image {:^20} has completed".format(image),end=" ")
        
    writer = pd.ExcelWriter("./200_image_jpeg_result.xlsx",engine='openpyxl')
    
    wb = load_workbook(writer.path)
    # wb = Workbook()
    writer.book = wb
    
    df = pd.DataFrame(image_psnr_dict)
    df.set_index("image")
    df.to_excel(writer,sheet_name="length {}".format(length),index=False)
    writer.save()
    writer.close()
    


